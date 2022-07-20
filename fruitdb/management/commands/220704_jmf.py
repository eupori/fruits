import requests
import urllib
import json
import csv
import os
import django
import datetime
import pandas as pd


import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.core.files import File
from django.db.models import Sum
from django.db.models import Q, F
from fruitdb.views.api.auth import get_username_from_uat, get_sat, search_user
from fruitdb.models import *
from dateutil.relativedelta import relativedelta
from rest_framework import serializers
from io import BytesIO as IO

class Settlement():
    def __init__(self):
        # self.today = datetime.date.today()
        self.today = datetime.datetime.strptime("2022-06-01", "%Y-%m-%d")
        self.url = "https://shopman.d-if.kr"
        self.file_name = "data/settlement.xlsx"
        self.contract_times = 12

        self.subs = JMFSubscription.objects.filter(
            start_datetime__year=self.today.year,
            start_datetime__month=self.today.month
            )

    # 구독상품 정산
    def subscription_product(self):
        products = GeneticTestProduct.objects.all()
        cols = ["구독 플랫폼", "구독 상품", "서비스 이용 회원", "약정기간 내 회원수", "약정기간 내 회원 결제 건수", "약정기간 내 회원 결제 금액", "약정기간 이후 회원수", "약정기간 이후 회원 결제 건수", "약정기간 이후 회원 결제 금액", "일시납 회원 결제 건수", "일시납 회원 결제 금액", "해지 회원수", "총결제 건수", "매출액", "수수료 정산"]
        person_data_list = []
        for product in products:
            logs = JMFSubscriptionPaymentLog.objects.filter(jmf_subscription__product=product, status="paid", order_date__year=self.today.year, order_date__month=self.today.month)
            sub_ids = logs.filter(jmf_subscription__product=product,).values_list("jmf_subscription", flat=True).distinct()
            subs = JMFSubscription.objects.filter(id__in=sub_ids)

            active_subs = JMFSubscription.objects.filter(is_active=True)
            deactive_subs = subs.filter(is_active=False, end_datetime__year=self.today.year, end_datetime__month=self.today.month)
            inround_subs = active_subs.filter(current_times__lte=self.contract_times, sub_type="0")
            outround_subs = active_subs.filter(Q(current_times__gt=self.contract_times, sub_type="0")|Q(sub_type="2"))
            lump_sum_subs = active_subs.filter(sub_type="1")
        
            in_round_logs = logs.filter(jmf_subscription__in=inround_subs, action="subscription")
            out_round_logs = logs.filter(jmf_subscription__in=outround_subs, action="subscription")
            lump_sum_logs = logs.filter(jmf_subscription__in=lump_sum_subs, action="subscription")

            data_dict = {}
            data_dict["구독 플랫폼"] = product.get_connection_client_display()
            data_dict["구독 상품"] = product.product_name
            data_dict["서비스 이용 회원"] = len(active_subs)
            data_dict["약정기간 내 회원수"] = len(inround_subs)
            data_dict["약정기간 내 회원 결제 건수"] = len(in_round_logs)
            data_dict["약정기간 내 회원 결제 금액"] = sum(in_round_logs.values_list("billing_price", flat=True))
            data_dict["약정기간 이후 회원수"] = len(outround_subs)
            data_dict["약정기간 이후 회원 결제 건수"] = len(out_round_logs)
            data_dict["약정기간 이후 회원 결제 금액"] = sum(out_round_logs.values_list("billing_price", flat=True))
            data_dict["일시납 회원 결제 건수"] = len(lump_sum_logs)
            data_dict["일시납 회원 결제 금액"] = sum(lump_sum_logs.values_list("billing_price", flat=True))
            data_dict["해지 회원수"] = len(deactive_subs)
            data_dict["총결제 건수"] = len(logs.filter(jmf_subscription__in=subs, action="subscription"))
            data_dict["매출액"] = sum(logs.filter(jmf_subscription__in=subs, action="subscription").values_list("billing_price", flat=True))

            w = sum(in_round_logs.values_list("billing_price", flat=True))*0.07
            x = sum(out_round_logs.values_list("billing_price", flat=True))*0.1
            z = sum(lump_sum_logs.values_list("billing_price", flat=True))*0.07
            data_dict["수수료 정산"] = w+x+z

            person_data_list.append(data_dict)

        df = pd.DataFrame.from_dict(person_data_list, orient='columns')
        person_data = {}
        person_data["구독 플랫폼"] = "총 합계"
        person_data["서비스 이용 회원"] = df["서비스 이용 회원"].sum()
        person_data["약정기간 내 회원수"] = df["약정기간 내 회원수"].sum()
        person_data["약정기간 내 회원 결제 건수"] = df["약정기간 내 회원 결제 건수"].sum()
        person_data["약정기간 내 회원 결제 금액"] = df["약정기간 내 회원 결제 금액"].sum()
        person_data["약정기간 이후 회원수"] = df["약정기간 이후 회원수"].sum()
        person_data["약정기간 이후 회원 결제 건수"] = df["약정기간 이후 회원 결제 건수"].sum()
        person_data["약정기간 이후 회원 결제 금액"] = df["약정기간 이후 회원 결제 금액"].sum()
        person_data["일시납 회원 결제 건수"] = df["일시납 회원 결제 건수"].sum()
        person_data["일시납 회원 결제 금액"] = df["일시납 회원 결제 금액"].sum()
        person_data["해지 회원수"] = df["해지 회원수"].sum()
        person_data["총결제 건수"] = df["총결제 건수"].sum()
        person_data["매출액"] = df["매출액"].sum()
        person_data["수수료 정산"] = df["수수료 정산"].sum()
        #"{:,}".format()

        df1 = df.append(person_data, ignore_index=True)

        return df1

    # 단품상품 정산
    def single_product(self):
        cols = ["단품상품", "단가", "판매 수량", "환불 건수", "판매액", "매출액(판매 - 환불)", "수수료 정산"]
        start_date = self.today
        end_date = self.today- datetime.timedelta(days=1)

        url = "https://shopman.d-if.kr"
        # url = "127.0.0.1:8001/product/fruits"
        response = requests.request("GET", f"{url}/product/fruits")
        products = json.loads(response.text)["data"]

        response = requests.request("GET", f"{url}/order/fruits?type=daily")
        order_list = json.loads(response.text)["data"]

        response = requests.request("GET", f"{url}/paymentlog/fruits?type=daily")
        paymentlog_list = json.loads(response.text)["data"]

        order_product_ids = list(set([item["product_id"] for item in order_list]))

        product_list = []
        for each in products:
            if each["id"] in order_product_ids:
                product_list.append(each)

        person_data_list = []

        for product in product_list:
            exclude_status_list = ["4", "5", "8"]
            include_status_list_refund = ["4", "5"]

            sell_list = []
            sell_price = 0
            refund_list = []
            refund_price = 0

            for payment in paymentlog_list:
                target_order = None
                for order in order_list:
                    if payment["order_id"] == order["id"]:
                        target_order = order
                if target_order:
                    if target_order["product_id"] == product["id"] and payment["method"] == "card" and target_order["price"] != 0 and payment["payment_name"] == "카드결제":
                        sell_list.append(target_order)
                    elif target_order["product_id"] == product["id"] and payment["method"] == None and payment["payment_name"] == "결제 취소":
                        refund_list.append(target_order)
                

            name = product["name"]
            price = product["price"]

            sell_cnt = len(sell_list)
            sell_price = sum([int(int(item["price"])/1.1) if item["product_id"]==53 else int(item["price"]) for item in sell_list])
            refund_cnt = len(refund_list)
            refund_price = sum([int(int(item["price"])/1.1) if item["product_id"]==53 else int(item["price"]) for item in refund_list])

            order_data = {}
            order_data["단품상품"] = name
            order_data["단가"] = price
            order_data["판매 수량"] = sell_cnt
            order_data["환불 건수"] = refund_cnt
            order_data["판매액"] = sell_price
            order_data["매출액(판매 - 환불)"] = sell_price - refund_price
            order_data["수수료 정산"] = int((sell_price - refund_price)/10)

            person_data_list.append(order_data)

        if len(person_data_list) > 0:
            df = pd.DataFrame.from_dict(person_data_list, orient='columns')
            order_data = {}
            order_data["단품상품"] = "총 합계"
            order_data["판매 수량"] = df["판매 수량"].sum()
            order_data["환불 건수"] = df["환불 건수"].sum()
            order_data["판매액"] = df["판매액"].sum()
            order_data["매출액(판매 - 환불)"] = df["매출액(판매 - 환불)"].sum()
            order_data["수수료 정산"] = df["수수료 정산"].sum()
            # #"{:,}".format()

            df2 = df.append(order_data, ignore_index=True)


    def payment_log(self):
        logs = JMFSubscriptionPaymentLog.objects.exclude(jmf_subscription=None).exclude(jmf_subscription__user_name=None).filter(order_date__year=self.today.year).filter(order_date__month=self.today.month)
        logs_qs = []
        max_kit_length = 0
        for item in logs:
            data = []
            data.append(item.order_date.date())
            data.append(item.order_date.strftime("%p %I:%M"))
            data.append(item.jmf_subscription.user_name if item.jmf_subscription.user_name else None)
            data.append(item.jmf_subscription.phone if item.jmf_subscription.phone else None)
            data.append(item.jmf_subscription.product.get_connection_client_display())
            data.append(item.jmf_subscription.order_id)
            data.append(item.jmf_subscription.product.product_code)
            data.append(item.jmf_subscription.product.product_name)
            data.append(item.get_action_display())
            data.append(item.current_times)
            data.append(item.billing_price)

            logs_qs.append(data)
        
        columns=["날짜", "시간", "고객", "전화번호", "거래처", "주문번호", "상품코드","상품명", "이벤트", "누적회차", "청구금액"]
        logs_df = pd.DataFrame(logs_qs, columns=columns)

        name_ids = list(set(logs_df["주문번호"].to_list()))
        name_ids.sort()
        df3 = pd.DataFrame(columns=columns)
        for name in name_ids:
            name_df = logs_df.groupby(["주문번호"]).get_group(name)
            #name_df = name_df.sort_values(["날짜", "시간"], axis=0)
            df3 = df3.append(name_df)
            df3 = df3.append(pd.Series(), ignore_index=True)
        df3 = df3.fillna("")

        return df3


    # 계약기간 내(중도 해지) 쿠폰 미사용(+3%)
    def coupon_unused_billing_in_active(self):
        bill_list = []
        genetic_test_products = GeneticTestProduct.objects.all()

        for genetic_test_product in genetic_test_products:
            coupon_unused_subs = self.subs.filter(
                    product=genetic_test_product,
                    is_active=False,
                    # 유전자 검사 상품만을 포함하기 위해 추가
                    coupon__isnull=False,
                    is_coupon_used=False,
                    current_times__lt=12,
                    )
            for sub in coupon_unused_subs:
                product_price = genetic_test_product.price
                count_payments = sub.current_times
                total_product_price = product_price * count_payments
                bill_price = (total_product_price * (3 * 0.7) / 100.0)

                bill_data = {}
                bill_data["client"] = genetic_test_product.connection_client
                bill_data["category"] = "계약기간 내(중도 해지) 쿠폰 미사용(+3%)"
                bill_data["product"] = genetic_test_product.product_name
                bill_data["price"] = bill_price
                bill_list.append(bill_data)


                bill = BillStatus.objects.filter(jmf_sub=sub)
                if len(bill) <= 0:
                    BillStatus.objects.create(
                            jmf_sub=sub,
                            user_name = sub.user_name,
                            status = 'coupon_bill_in_active',
                            price = bill_price
                        )
        if len(bill_list) > 0:
            coupon_not_used_billing = pd.DataFrame.from_dict(bill_list, orient='columns')
            return coupon_not_used_billing
        else:
            bill_data = {}
            bill_data["client"] = ""
            bill_data["category"] = "계약기간 내(중도 해지) 쿠폰 미사용(+3%)"
            bill_data["product"] = "None"
            bill_data["price"] = 0
            bill_list.append(bill_data)
            coupon_not_used_billing = pd.DataFrame.from_dict(bill_list, orient='columns')
            return coupon_not_used_billing


    # 계약기간 내(중도 해지) 쿠폰 사용(-3%)
    def coupon_used_refund_in_active(self):
        refund_list = []
        genetic_test_products = GeneticTestProduct.objects.all()

        for genetic_test_product in genetic_test_products:
            product_price = genetic_test_product.price
            coupon_used_subs = self.subs.filter(
                    product=genetic_test_product,
                    is_active=False,
                    is_coupon_used=True,
                    current_times__lt=12,
                    )

            for sub in coupon_used_subs:
                count_payments = sub.current_times
                total_product_price = product_price * count_payments
                bill = BillStatus.objects.filter(
                            jmf_sub=sub,
                            status='coupon_bill_in_active',    
                            )
                if len(bill) > 0:
                    bill = bill[0]
                    refund_data = {}
                    refund_data["client"] = genetic_test_product.connection_client
                    refund_data["category"] = "계약기간 내(중도 해지) 쿠폰 사용(-3%)"
                    refund_data['product'] = genetic_test_product.product_name
                    refund_data['price'] = -(bill.price)
                    refund_list.append(refund_data)

                    bill.status = 'coupon_refund'
                    bill.save()

        if len(refund_list) > 0:
            coupon_used_refund = pd.DataFrame.from_dict(refund_list, orient='columns')
            return coupon_used_refund
        else:
            refund_data = {}
            refund_data["client"] = ""
            refund_data["category"] = "계약기간 내(중도 해지) 쿠폰 사용(-3%)"
            refund_data['product'] = "None"
            refund_data['price'] = 0
            refund_list.append(refund_data)
            coupon_used_refund = pd.DataFrame.from_dict(refund_list, orient='columns')
            return coupon_used_refund

    '''
    - 약정은 기본 12회- 사용자가 비용을 지불 했을 때, 1회차 증가
    - 약정기간 만료 사용자란, 약정회차(12회)를 지난 사용자를 말함
    - 청구서는 지난달 결제를 기준으로 하기에, current_times__gt=12 일 경우, 약정만료 사용자로 설정
    (지난달에 12회차 결제를 하고, 회차가 끝날시점에 청구서를 작성하기 때문)
    '''
    # 계약서 4조 4번 사항
    def coupon_unused_billing(self):
        bill_list = []
        genetic_test_products = GeneticTestProduct.objects.all()

        for genetic_test_product in genetic_test_products:
            coupon_unused_subs = self.subs.filter(
                    product=genetic_test_product,
                    #is_active=False,
                    # 유전자 검사 상품만을 포함하기 위해 추가
                    coupon__isnull=False,
                    is_coupon_used=False,
                    current_times__gt=12,
                    )

            for sub in coupon_unused_subs:
                product_price = genetic_test_product.price
                count_payments = self.contract_times
                total_product_price = product_price * count_payments
                bill_price = (total_product_price * (3 * 0.7) / 100.0)

                bill_data = {}
                bill_data["client"] = genetic_test_product.connection_client
                bill_data["category"] = "계약기간 만료 회차 쿠폰 미사용(+3%)"
                bill_data["product"] = genetic_test_product.product_name
                bill_data["price"] = bill_price
                bill_list.append(bill_data)


                bill = BillStatus.objects.filter(jmf_sub=sub)
                if len(bill) <= 0:
                    BillStatus.objects.create(
                            jmf_sub=sub,
                            user_name = sub.user_name,
                            status = 'coupon_bill',
                            price = bill_price
                        )

        if len(bill_list) > 0:
            coupon_not_used_billing = pd.DataFrame.from_dict(bill_list, orient='columns')
            return coupon_not_used_billing
        else:
            bill_data = {}
            bill_data["client"] = ""
            bill_data["category"] = "계약기간 만료 회차 쿠폰 미사용(+3%)"
            bill_data["product"] = "None"
            bill_data["price"] = 0
            bill_list.append(bill_data)
            coupon_not_used_billing = pd.DataFrame.from_dict(bill_list, orient='columns')
            return coupon_not_used_billing


    # 계약서 4조 5번 사항
    def coupon_used_refund(self):
        refund_list = []
        genetic_test_products = GeneticTestProduct.objects.all()

        for genetic_test_product in genetic_test_products:
            product_price = genetic_test_product.price
            coupon_used_subs = self.subs.filter(
                    product=genetic_test_product,
                    #is_active=False,
                    is_coupon_used=True,
                    current_times__gt=12,
                    )

            for sub in coupon_used_subs:
                count_payments = self.contract_times
                total_product_price = product_price * count_payments
                bill = BillStatus.objects.filter(
                            jmf_sub=sub,
                            status='coupon_bill',    
                            )
                if len(bill) > 0:
                    bill = bill[0]
                    refund_data = {}
                    refund_data["client"] = genetic_test_product.connection_client
                    refund_data["category"] = "계약기간 이후 쿠폰 사용(-3%)"
                    refund_data['product'] = genetic_test_product.product_name
                    refund_data['price'] = -(bill.price)
                    refund_list.append(refund_data)

                    bill.status = 'coupon_refund'
                    bill.save()

        if len(refund_list) > 0:
            coupon_used_refund = pd.DataFrame.from_dict(refund_list, orient='columns')
            return coupon_used_refund
        else:
            refund_data = {}
            refund_data["client"] = ""
            refund_data["category"] = "계약기간 이후 쿠폰 사용(-3%)"
            refund_data['product'] = "None"
            refund_data['price'] = 0
            refund_list.append(refund_data)
            coupon_used_refund = pd.DataFrame.from_dict(refund_list, orient='columns')
            return coupon_used_refund


    # 계약서 4조 6번 사항
    def kit_unused_billing(self):
        bill_list = []
        genetic_test_products = GeneticTestProduct.objects.all()

        for genetic_test_product in genetic_test_products:
            product_price = genetic_test_product.price
            kit_unused_subs = self.subs.filter(
                    product=genetic_test_product,
                    #is_active=False,
                    is_coupon_used=True,
                    current_times__gt=12,
                    kit_receive_datetime__isnull=True,
                    )
            for sub in kit_unused_subs:
                count_payments = self.contract_times
                total_product_price = product_price * count_payments
                bill_price = (total_product_price * (1.5 * 0.7) / 100.0)

                bill_data = {}
                bill_data["client"] = genetic_test_product.connection_client
                bill_data["category"] = "키트 미사용(+1.5%)"
                bill_data["product"] = genetic_test_product.product_name
                bill_data["price"] = bill_price
                bill_list.append(bill_data)


                bill = BillStatus.objects.filter(jmf_sub=sub)
                if len(bill) <= 0:
                    BillStatus.objects.create(
                            jmf_sub=sub,
                            user_name = sub.user_name,
                            status = 'kit_bill',
                            price = bill_price
                        )

        if len(bill_list) > 0:
            kit_not_used_billing = pd.DataFrame.from_dict(bill_list, orient='columns')
            return kit_not_used_billing
        else:
            bill_data = {}
            bill_data["client"] = ""
            bill_data["category"] = "키트 미사용(+1.5%)"
            bill_data["product"] = "None"
            bill_data["price"] = 0
            bill_list.append(bill_data)
            kit_not_used_billing = pd.DataFrame.from_dict(bill_list, orient='columns')
            return kit_not_used_billing


    # 계약서 4조 7번 사항
    def kit_used_refund(self):
        refund_list = []
        genetic_test_products = GeneticTestProduct.objects.all()

        for genetic_test_product in genetic_test_products:
            product_price = genetic_test_product.price
            kit_used_subs = self.subs.filter(
                    product=genetic_test_product,
                    #is_active=False,
                    is_coupon_used=True,
                    current_times__gt=12,
                    kit_receive_datetime__isnull=False,
                    )

            for sub in kit_used_subs:
                count_payments = self.contract_times
                total_product_price = product_price * count_payments
                bill = BillStatus.objects.filter(
                            jmf_sub=sub,
                            status='kit_bill',    
                            )
                if len(bill) > 0:
                    bill = bill[0]
                    refund_data = {}
                    refund_data["client"] = genetic_test_product.connection_client
                    refund_data["category"] = "키트 사용(-1.5%)"
                    refund_data['product'] = genetic_test_product.product_name
                    refund_data['price'] = -(bill.price)
                    refund_list.append(refund_data)

                    bill.status = 'kit_refund'
                    bill.save()

        if len(refund_list) > 0:
            coupon_used_refund = pd.DataFrame.from_dict(refund_list, orient='columns')
            return coupon_used_refund
        else:
            refund_data = {}
            refund_data["client"] = ""
            refund_data["category"] = "키트 사용(-1.5%)"
            refund_data['product'] = "None"
            refund_data['price'] = 0
            refund_list.append(refund_data)
            coupon_used_refund = pd.DataFrame.from_dict(refund_list, orient='columns')
            return coupon_used_refund

    # 계약서 4조 8번 사항
    def kit_refund_billing(self):
        bill_list = []
        genetic_test_products = GeneticTestProduct.objects.all()

        for genetic_test_product in genetic_test_products:
            product_price = genetic_test_product.price
            kit_return_subs = self.subs.filter(
                    product=genetic_test_product,
                    #is_active=False,
                    is_coupon_used=True,
                    current_times__gt=12,
                    kit_receive_datetime__isnull=True,
                    )
            '''
            구독정보(JMFSubscription)의 주문 정보와 
            유전자검사(GeneticTest)의 주문번호가 일치한다고함. (by 이채민 주임)
            구독 정보의 주문번호를 활용하여 GeneticTest 의 정보를 가져와서,
            GeneticTestLog 검색에 사용(키트 반송 정보 확인용도)
            '''
            for sub in kit_return_subs:
                genetic_test = GeneticTest.objects.filter(
                        order_id=sub.order_id,
                        )
                # 주문번호(order_id)는 유일할 것이기 때문에,
                # genetic_test[0] 첫 번째 값만 사용.
                logs = None
                if len(genetic_test) > 0:
                    logs = GeneticTestLog.objects.filter(
                            genetic_test=genetic_test[0]
                            )

                count_payments = self.contract_times
                total_product_price = product_price * count_payments
                bill_price = None
                send_date = None
                
                if logs is not None: 
                    for log in logs:
                        if log.action == 'kit_send':
                            send_date = log.order_date
                            after_two_week = send_date + datetime.timedelta(weeks=2)
                        if log.action == 'normal_kit_return':
                            if (log.order_date > after_two_week):
                                bill_price = (total_product_price * (1.5 * 0.7) / 100.0)
                                bill_data = {}
                                bill_data["client"] = genetic_test_product.connection_client
                                bill_data["category"] = "2주 후 키트 반품(+1.5%)"
                                bill_data["product"] = genetic_test_product.product_name
                                bill_data["price"] = bill_price
                                bill_list.append(bill_data)
                            else:
                                bill_price = (total_product_price * (3.0 * 0.7) / 100.0)
                                bill_data = {}
                                bill_data["client"] = genetic_test_product.connection_client
                                bill_data["category"] = "정상 키트 반품(+3%)"
                                bill_data["product"] = genetic_test_product.product_name
                                bill_data["price"] = bill_price
                                bill_list.append(bill_data)
                        if log.action == 'damaged_kit_return':
                                bill_price = (total_product_price * (1.5 * 0.7) / 100.0)
                                bill_data = {}
                                bill_data["client"] = genetic_test_product.connection_client
                                bill_data["category"] = "파손키트 반품(+1.5%)"
                                bill_data["product"] = genetic_test_product.product_name
                                bill_data["price"] = bill_price
                                bill_list.append(bill_data)

        if len(bill_list) > 0:
            kit_return = pd.DataFrame.from_dict(bill_list, orient='columns')
            return kit_return
        else:
            bill_data = {}
            bill_data["client"] = ""
            bill_data["category"] = "2주 후 키트 반품(+1.5%)"
            bill_data["product"] = "None"
            bill_data["price"] = 0
            bill_list.append(bill_data)

            bill_data = {}
            bill_data["client"] = ""
            bill_data["category"] = "정상 키트 반품(+3%)"
            bill_data["product"] = "None"
            bill_data["price"] = 0
            bill_list.append(bill_data)

            bill_data = {}
            bill_data["client"] = ""
            bill_data["category"] = "파손키트 반품(+1.5%)"
            bill_data["product"] = "None"
            bill_data["price"] = 0
            bill_list.append(bill_data)

            kit_return = pd.DataFrame.from_dict(bill_list, orient='columns')
            return kit_return

    def get_key(self, dic, val):
        for key, value in dic.items():
            if val == value:
                return key
        return "There is no such Key"

    # 정산 엑셀 시트에 저장
    '''
    CLIENT_COICES = (
        ("1", _("카카오")),
        ("2", _("고도몰")),
        ("999", _("과일궁합")),
    )
    '''
    def save_settlement(self):
        genetic_test_product = GeneticTestProduct()
        client_choices = genetic_test_product.CLIENT_COICES
        coupon_unused_billing_in_active = self.coupon_unused_billing_in_active()
        coupon_used_refund_in_active = self.coupon_used_refund_in_active()
        coupon_unused_billing = self.coupon_unused_billing()
        coupon_used_refund = self.coupon_used_refund()
        kit_unused_billing = self.kit_unused_billing()
        kit_used_refund = self.kit_used_refund()
        kit_refund_billing = self.kit_refund_billing()
        #tmp_df = pd.DataFrame(columns=["client", "category", "product", "price"])
        #kit_refund_billing = self.kit_refund_billing() if self.kit_refund_billing() else tmp_df

        settlements = pd.concat([coupon_unused_billing_in_active,
                                coupon_used_refund_in_active,
                                coupon_unused_billing,
                                coupon_used_refund,
                                kit_unused_billing,
                                kit_used_refund,
                                kit_refund_billing])

        ret = dict(map(reversed, client_choices))
        clients = set(settlements["client"])
        result = None
        for client in clients:
            data = {}
            if client in ret.values():
                client_name = self.get_key(ret, client)
                data["구독 플랫폼"] = client_name
            else:
                data["구독 플랫폼"] = client
            condition = settlements["client"] == client
            client_filter = settlements[condition]
            cols = set(client_filter["product"])
            for col in cols:
                settlements_list = []
                condition = client_filter["product"] == col
                product_filter = client_filter[condition]
                categorys = set(product_filter["category"])
                total_count = 0
                total_price = 0
                for category in categorys:
                    category_filter = product_filter[(product_filter.category == category)]
                    datas = category_filter
                    data["상품명"] = col
                    data[datas["category"].unique()[0]] = datas["category"].count()
                    total_count += datas["category"].count()
                    total_price += datas["price"].sum()
                data["총 발생 건 수"] = total_count
                data["기타 수수료"] = round(total_price)
                settlements_list.append(data)
                settlements_df = pd.DataFrame.from_dict(settlements_list, orient='columns')
                result = pd.concat([result, settlements_df])
                result.fillna(0, inplace=True)
        index = []
        for i in range(0, len(result)):
            index.append(i)

        result.index = index
        drop_col = result[result["상품명"] == "None"].index
        result.drop(drop_col, inplace=True)
        result = result[[
                        "구독 플랫폼",
                        "상품명",
                        "계약기간 내(중도 해지) 쿠폰 미사용(+3%)",
                        "계약기간 내(중도 해지) 쿠폰 사용(-3%)",
                        "계약기간 만료 회차 쿠폰 미사용(+3%)",
                        "계약기간 이후 쿠폰 사용(-3%)",
                        "키트 미사용(+1.5%)",
                        "키트 사용(-1.5%)",
                        "정상 키트 반품(+3%)",
                        "2주 후 키트 반품(+1.5%)",
                        "파손키트 반품(+1.5%)",
                        "총 발생 건 수",
                        "기타 수수료",
                        ]]
        if result is not None and not result.empty:
            return result
        else:
            return None


def get_settlement(first_datetime, last_datetime):
    today = datetime.datetime.today()
    products = GeneticTestProduct.objects.all()

    settlement = Settlement()
    df1 = settlement.subscription_product()
    df2 = settlement.single_product()
    df3 = settlement.payment_log()
    df4 = settlement.save_settlement()

    wb = openpyxl.load_workbook("data/jmf_settlement_sheet_daily.xlsx")
    wb.save("data/tmp_jmf.xlsx")

    excel_writer = pd.ExcelWriter("data/tmp_jmf.xlsx", engine='openpyxl', mode='a')

    df1.to_excel(excel_writer, sheet_name="구독상품", index=False)
    if df2 is not None and not df2.empty:
        df2.to_excel(excel_writer, sheet_name="단품상품", index=False)
    df3.to_excel(excel_writer, sheet_name="결제내역", index=False)
    if df4 is not None and not df4.empty:
        df4.to_excel(excel_writer, sheet_name="기타 수수료", index=False)

    excel_writer.save()

    wb = openpyxl.load_workbook("data/tmp_jmf.xlsx")
    wb.active
    for col in range(1, 19):
        wb["구독상품"].column_dimensions[get_column_letter(col)].width = 15
        if df2 is not None and not df2.empty:
            wb["단품상품"].column_dimensions[get_column_letter(col)].width = 15
        wb["결제내역"].column_dimensions[get_column_letter(col)].width = 15
    wb["구독상품"].column_dimensions[get_column_letter(2)].width = 50
    if df2 is not None and not df2.empty:
        wb["단품상품"].column_dimensions[get_column_letter(1)].width = 30
    wb.save("data/tmp_jmf.xlsx")

    result = JMFSubscriptionLogResult()
    today = datetime.datetime.now()

    with open('data/tmp_jmf.xlsx', 'rb') as excel:
            result.result_file_excel.save('jmf_settlement_'+today.strftime('%Y-%m-%d %H:%M:%S')+'.xlsx', File(excel))
            result.save()


# 매달 말일 23:59:59에 스크립트 동작
class Command(BaseCommand):
    help = '인실리코젠-진맛과 정산'

    def handle(self, *args, **options):
        today = datetime.datetime.today()
        this_month_first = datetime.datetime(today.year, today.month, 1)
        next_month = datetime.datetime(today.year, today.month, 1) + relativedelta(months=1)
        this_month_last = next_month + relativedelta(seconds=-1)

        print('이번달 첫일: ' + this_month_first.strftime('%Y-%m-%d %H:%M:%S'))
        print('이번달 말일: ' + this_month_last.strftime('%Y-%m-%d %H:%M:%S'))

        first_datetime = this_month_first.strftime('%Y-%m-%d %H:%M:%S')
        last_datetime = this_month_last.strftime('%Y-%m-%d %H:%M:%S')
        last_date = this_month_last.strftime('%Y-%m-%d')
        get_settlement(first_datetime, last_datetime)
