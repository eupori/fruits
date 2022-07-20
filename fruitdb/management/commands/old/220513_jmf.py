import requests
import urllib
import json
import csv
import os
import django
import datetime
import pandas as pd


import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.core.files import File
from django.db.models import Sum
from django.db.models import Q
from fruitdb.views.api.auth import get_username_from_uat, get_sat, search_user
from fruitdb.models import *
from dateutil.relativedelta import relativedelta
from rest_framework import serializers
from io import BytesIO as IO


def get_settlement(first_datetime, last_datetime):

    today = datetime.datetime.strptime("2022-05-16", '%Y-%m-%d')
    today = datetime.datetime.today()

    first_date=datetime.datetime.strptime(first_datetime, '%Y-%m-%d %H:%M:%S')
    last_date=datetime.datetime.strptime(last_datetime, '%Y-%m-%d %H:%M:%S')

    cols = ["구독 플랫폼", "구독 상품", "구독중 회원", "회차 내 회원(12회차 이전)", "회차 내 회원 결제 건수", "회차 내 회원 결제 금액", "회차 완료 회원(12회차 이후)", "회차 완료 회원 결제 건수", "회차 완료 회원 결제 금액", "쿠폰 미사용 회원", "쿠폰 미사용 회원 결제 건수", "쿠폰 미사용 회원 결제 금액", "해지 회원", "해지 회원 결제 건수", "해지 회원 결제 금액", "총 결제 건 수", "매출액", "수수료 정산"]
    start_date = today
    end_date = today.date() - datetime.timedelta(days=1)
    
    genetic_test_products = GeneticTestProduct.objects.all()
    subs = JMFSubscription.objects.filter(start_datetime__year=today.year, start_datetime__month=today.month, start_datetime__day=today.day)

    person_data_list = []
    for genetic_test_product in genetic_test_products:
        in_round_subs = subs.filter(product=genetic_test_product, is_active=True, is_coupon_used=True, current_times__lte=12)
        out_round_subs = subs.filter(product=genetic_test_product, is_active=True, is_coupon_used=True, current_times__gt=12)
        coupon_unused_subs = subs.filter(product=genetic_test_product, is_active=True, is_coupon_used=False)
        cancel_subs =  subs.filter(product=genetic_test_product, is_active=False, end_datetime__year=today.year, end_datetime__month=today.month)

        in_round_logs = JMFSubscriptionPaymentLog.objects.filter(jmf_subscription__in=in_round_subs, action="subscription", order_date__year=today.year, order_date__month=today.month)
        out_round_logs = JMFSubscriptionPaymentLog.objects.filter(jmf_subscription__in=out_round_subs, action="subscription", order_date__year=today.year, order_date__month=today.month)
        coupon_unused_logs = JMFSubscriptionPaymentLog.objects.filter(jmf_subscription__in=coupon_unused_subs, action="subscription", order_date__year=today.year, order_date__month=today.month)
        cancel_logs = JMFSubscriptionPaymentLog.objects.filter(jmf_subscription__in=cancel_subs, action="subscription", order_date__year=today.year, order_date__month=today.month)

        person_data = {}
        person_data["구독 플랫폼"] = genetic_test_product.get_connection_client_display()
        person_data["구독 상품"] = genetic_test_product.product_name
        person_data["구독중 회원"] = len(subs.filter(product=genetic_test_product, is_active=True))
        person_data["회차 내 회원(12회차 이전)"] = len(in_round_subs)
        person_data["회차 내 회원 결제 건수"] = len(in_round_logs)
        person_data["회차 내 회원 결제 금액"] = sum(in_round_logs.values_list("billing_price", flat=True))
        person_data["회차 완료 회원(12회차 이후)"] = len(out_round_subs)
        person_data["회차 완료 회원 결제 건수"] = len(out_round_logs)
        person_data["회차 완료 회원 결제 금액"] = sum(out_round_logs.values_list("billing_price", flat=True))
        person_data["쿠폰 미사용 회원"] = len(coupon_unused_subs)
        person_data["쿠폰 미사용 회원 결제 건수"] = len(coupon_unused_logs)
        person_data["쿠폰 미사용 회원 결제 금액"] = sum(coupon_unused_logs.values_list("billing_price", flat=True))
        person_data["해지 회원"] = len(cancel_subs)
        person_data["해지 회원 결제 건수"] = len(cancel_logs)
        person_data["해지 회원 결제 금액"] = sum(cancel_logs.values_list("billing_price", flat=True))
        person_data["총 결제 건 수"] = len(in_round_logs) + len(out_round_logs) + len(coupon_unused_logs) + len(cancel_logs)
        person_data["매출액"] = sum(in_round_logs.values_list("billing_price", flat=True)) + sum(out_round_logs.values_list("billing_price", flat=True)) + sum(coupon_unused_logs.values_list("billing_price", flat=True)) + sum(cancel_logs.values_list("billing_price", flat=True))

        w = sum(in_round_logs.values_list("billing_price", flat=True))*0.07
        x = sum(out_round_logs.values_list("billing_price", flat=True))*0.1
        y = sum(coupon_unused_logs.values_list("billing_price", flat=True))*0.07
        z = sum(cancel_logs.values_list("billing_price", flat=True))*0.07 + sum([ item.current_times*item.product.billing_price for item in cancel_subs ])*0.03
        person_data["수수료 정산"] = w+x+y+z

        person_data_list.append(person_data)

    df = pd.DataFrame.from_dict(person_data_list, orient='columns')
    person_data = {}
    person_data["구독 플랫폼"] = "총 합계"
    person_data["구독중 회원"] = df["구독중 회원"].sum()
    person_data["회차 내 회원(12회차 이전)"] = df["회차 내 회원(12회차 이전)"].sum()
    person_data["회차 내 회원 결제 건수"] = df["회차 내 회원 결제 건수"].sum()
    person_data["회차 내 회원 결제 금액"] = df["회차 내 회원 결제 금액"].sum()
    person_data["회차 완료 회원(12회차 이후)"] = df["회차 완료 회원(12회차 이후)"].sum()
    person_data["회차 완료 회원 결제 건수"] = df["회차 완료 회원 결제 건수"].sum()
    person_data["회차 완료 회원 결제 금액"] = df["회차 완료 회원 결제 금액"].sum()
    person_data["쿠폰 미사용 회원"] = df["쿠폰 미사용 회원"].sum()
    person_data["쿠폰 미사용 회원 결제 건수"] = df["쿠폰 미사용 회원 결제 건수"].sum()
    person_data["쿠폰 미사용 회원 결제 금액"] = df["쿠폰 미사용 회원 결제 금액"].sum()
    person_data["해지 회원"] = df["해지 회원"].sum()
    person_data["해지 회원 결제 건수"] = df["해지 회원 결제 건수"].sum()
    person_data["총 결제 건 수"] = df["총 결제 건 수"].sum()
    person_data["매출액"] = df["매출액"].sum()
    person_data["수수료 정산"] = df["수수료 정산"].sum()
    #"{:,}".format()

    df1 = df.append(person_data, ignore_index=True)

    df1.to_excel("jmf.xlsx", sheet_name="5월 과일궁합 정산(구독 상품)", index=False)

    wb = openpyxl.load_workbook("jmf.xlsx")
    wb.active
    for col in range(1, 19):
        wb["5월 과일궁합 정산(구독 상품)"].column_dimensions[get_column_letter(col)].width = 25
    wb["5월 과일궁합 정산(구독 상품)"].column_dimensions[get_column_letter(2)].width = 50
    wb["5월 과일궁합 정산(구독 상품)"].column_dimensions[get_column_letter(1)].width = 15
    wb["5월 과일궁합 정산(구독 상품)"].column_dimensions[get_column_letter(16)].width = 15
    wb["5월 과일궁합 정산(구독 상품)"].column_dimensions[get_column_letter(17)].width = 15
    wb["5월 과일궁합 정산(구독 상품)"].column_dimensions[get_column_letter(18)].width = 15

    wb.save("jmf.xlsx")

    cols = ["단품상품", "단가", "판매 수량", "환불 건수", "판매액", "매출액(판매 - 환불)", "수수료 정산"]
    start_date = today.date()
    end_date = today.date() - datetime.timedelta(days=1)

    url = "https://shopman.d-if.kr"
    # url = "127.0.0.1:8001/product/fruits"
    response = requests.request("GET", f"{url}/product/fruits")
    products = json.loads(response.text)["data"]

    response = requests.request("GET", f"{url}/order/fruits?type=daily")
    order_list = json.loads(response.text)["data"]

    response = requests.request("GET", f"{url}/paymentlog/fruits?type=daily")
    paymentlog_list = json.loads(response.text)["data"]

    order_product_ids = list(set([item["product_id"] for item in order_list]))

    product_list = []
    for each in products:
        if each["id"] in order_product_ids:
            product_list.append(each)

    person_data_list = []

    for product in product_list:
        exclude_status_list = ["4", "5", "8"]
        include_status_list_refund = ["4", "5"]

        sell_list = []
        sell_price = 0
        refund_list = []
        refund_price = 0

        for payment in paymentlog_list:
            target_order = None
            for order in order_list:
                if payment["order_id"] == order["id"]:
                    target_order = order
            if target_order:
                if target_order["product_id"] == product["id"] and payment["method"] == "card" and target_order["price"] != 0 and payment["payment_name"] == "카드결제":
                    sell_list.append(target_order)
                elif target_order["product_id"] == product["id"] and payment["method"] == None and payment["payment_name"] == "결제 취소":
                    refund_list.append(target_order)
            

        name = product["name"]
        price = product["price"]

        sell_cnt = len(sell_list)
        sell_price = sum([int(int(item["price"])/1.1) if item["product_id"]==53 else int(item["price"]) for item in sell_list])
        refund_cnt = len(refund_list)
        refund_price = sum([int(int(item["price"])/1.1) if item["product_id"]==53 else int(item["price"]) for item in refund_list])

        order_data = {}
        order_data["단품상품"] = name
        order_data["단가"] = price
        order_data["판매 수량"] = sell_cnt
        order_data["환불 건수"] = refund_cnt
        order_data["판매액"] = sell_price
        order_data["매출액(판매 - 환불)"] = sell_price - refund_price
        order_data["수수료 정산"] = int((sell_price - refund_price)/10)

        person_data_list.append(order_data)

    if len(person_data_list) > 0:
        df = pd.DataFrame.from_dict(person_data_list, orient='columns')
        order_data = {}
        order_data["단품상품"] = "총 합계"
        order_data["판매 수량"] = df["판매 수량"].sum()
        order_data["환불 건수"] = df["환불 건수"].sum()
        order_data["판매액"] = df["판매액"].sum()
        order_data["매출액(판매 - 환불)"] = df["매출액(판매 - 환불)"].sum()
        order_data["수수료 정산"] = df["수수료 정산"].sum()
        # #"{:,}".format()

        df2 = df.append(order_data, ignore_index=True)

        df2.to_excel("jmf_product.xlsx", sheet_name="5월 과일궁합 정산(단품 상품)", index=False)

        wb = openpyxl.load_workbook("jmf_product.xlsx")
        wb.active
        for col in range(1, 19):
            wb["5월 과일궁합 정산(단품 상품)"].column_dimensions[get_column_letter(col)].width = 15
        wb["5월 과일궁합 정산(단품 상품)"].column_dimensions[get_column_letter(1)].width = 25
        wb.save("jmf_product.xlsx")

    logs = JMFSubscriptionPaymentLog.objects.exclude(jmf_subscription=None).exclude(jmf_subscription__customer=None)
    logs_qs = []
    max_kit_length = 0
    for item in logs:
        data = []
        data.append(item.order_date.date())
        data.append(item.order_date.strftime("%p %I:%M"))
        data.append(item.jmf_subscription.customer.name if item.jmf_subscription.customer else None)
        data.append(item.jmf_subscription.customer.username if item.jmf_subscription.customer else None)
        data.append(item.jmf_subscription.phone.as_national if item.jmf_subscription.phone else None)
        data.append(item.jmf_subscription.product.get_connection_client_display())
        data.append(item.jmf_subscription.product.product_code)
        data.append(item.jmf_subscription.product.product_name)
        data.append(item.get_action_display())
        data.append(item.current_times)

        logs_qs.append(data)
    
    columns=["날짜", "시간", "고객", "아이디", "전화번호", "거래처", "상품코드","상품명", "이벤트", "누적회차"]
    logs_df = pd.DataFrame(logs_qs, columns=columns)

    user_ids = list(set(logs_df["아이디"].to_list()))
    df3 = pd.DataFrame(columns=columns)
    for user in user_ids:
        user_df = logs_df.groupby(["아이디"]).get_group(user)
        user_df = user_df.sort_values(["날짜", "시간"], axis=0)
        df3 = df3.append(user_df)
        df3 = df3.append(pd.Series(), ignore_index=True)
    df3 = df3.fillna("")

    df3.to_excel("jmf_log.xlsx", sheet_name="5월 과일궁합 정산(단품 상품)", index=False)

    wb = openpyxl.load_workbook("jmf_log.xlsx")
    wb.active
    for col in range(1, 19):
        wb["5월 과일궁합 정산(단품 상품)"].column_dimensions[get_column_letter(col)].width = 15
    wb["5월 과일궁합 정산(단품 상품)"].column_dimensions[get_column_letter(1)].width = 25
    wb.save("jmf_log.xlsx")

    wb = openpyxl.load_workbook("data/jmf_settlement_sheet_daily.xlsx")
    wb.save("data/tmp_jmf.xlsx")

    excel_writer = pd.ExcelWriter("data/tmp_jmf.xlsx", engine='openpyxl', mode='a')
    df1.to_excel(excel_writer, sheet_name="구독상품", index=False)
    if len(person_data_list) > 0:
        df2.to_excel(excel_writer, sheet_name="단품상품", index=False)
    df3.to_excel(excel_writer, sheet_name="결제내역", index=False)
    excel_writer.save()


# 매달 말일 23:59:59에 스크립트 동작
class Command(BaseCommand):
    help = '인실리코젠-진맛과 정산'

    def handle(self, *args, **options):
        today = datetime.datetime.today()
        this_month_first = datetime.datetime(today.year, today.month, 1)
        next_month = datetime.datetime(today.year, today.month, 1) + relativedelta(months=1)
        this_month_last = next_month + relativedelta(seconds=-1)

        print('이번달 첫일: ' + this_month_first.strftime('%Y-%m-%d %H:%M:%S'))
        print('이번달 말일: ' + this_month_last.strftime('%Y-%m-%d %H:%M:%S'))

        first_datetime = this_month_first.strftime('%Y-%m-%d %H:%M:%S')
        last_datetime = this_month_last.strftime('%Y-%m-%d %H:%M:%S')
        last_date = this_month_last.strftime('%Y-%m-%d')
        
        # if today.strftime('%Y-%m-%d') == last_date:
        #     get_settlement(first_datetime, last_datetime, macrogen_data_list)
        get_settlement(first_datetime, last_datetime)

        # last_month_first = datetime(today.year, today.month, 1) + relativedelta(months=-1)
        # last_month_last = datetime(today.year, today.month, 1) + relativedelta(seconds=-1)

        # print('지난달 첫일: ' + last_month_first.strftime('%Y-%m-%d %H:%M:%S'))
        # print('지난달 말일: ' + last_month_last.strftime('%Y-%m-%d %H:%M:%S'))