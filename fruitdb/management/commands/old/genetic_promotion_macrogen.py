import requests
import urllib
import json
import csv
import os
import django
import datetime
import pandas as pd


import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.core.files import File
from django.db.models import Sum
from django.db.models import Q
from fruitdb.views.api.auth import get_username_from_uat, get_sat, search_user
from fruitdb.models import *
from dateutil.relativedelta import relativedelta
from rest_framework import serializers
from io import BytesIO as IO


def get_settlement(first_datetime, last_datetime):

    first_date=datetime.datetime.strptime(first_datetime, '%Y-%m-%d %H:%M:%S')
    last_date=datetime.datetime.strptime(last_datetime, '%Y-%m-%d %H:%M:%S')

    
    # settlements = GeneticPromotionLog.objects.filter(order_date__gte=first_date, order_date__lte=last_date)
    # coupons = settlements.values_list("genetic_promotion__coupon", flat=True).distinct()
    genetic_tests = GeneticTest.objects.filter(end_datetime=None, is_active=True)

    # person_data_list = []
    # for coupon in coupons:
    #     person_objects = settlements.filter(genetic_promotion__coupon=coupon)
    #     person_data = person_objects.aggregate(정산금액=Sum("billing_price"))
    #     person_object = person_objects.last()
    #     person_data["쿠폰번호"] = coupon
    #     person_data["매출 항목명"] = person_object.item_name
    #     person_data["거래처"] = person_object.genetic_promotion.promotion_product.connection_client
    #     person_data["상품코드"] = person_object.genetic_promotion.promotion_product.product_code
    #     person_data["상품명"] = person_object.genetic_promotion.promotion_product.product_name
    #     person_data["고객명"] = person_object.genetic_promotion.customer_name
    #     person_data["아이디"] = person_object.genetic_promotion.customer_id
    #     person_data["이동전화번호"] = person_object.genetic_promotion.phone_number
    #     person_data["키트번호"] = person_object.promotion_kit.kit_code if person_object.promotion_kit else ""
    #     person_data["가입일"] = person_object.genetic_promotion.sign_datetime
    #     person_data["당월 배송 회차"] = person_object.current_month_times
    #     person_data["남은 회차"] = person_object.left_times
    #     person_data["월별 누적"] = person_object.accumulate_times
    #     person_data["해지일"] = person_object.genetic_promotion.termination_datetime if person_object.genetic_promotion.termination_datetime else ""
    #     person_data_list.append(person_data)

    # df = pd.DataFrame.from_dict(person_data_list, orient='columns')

    person_data_list = []
    for genetic_test in genetic_tests:
        print(genetic_test.kit_current_times)
        if genetic_test.kit_current_times ==0:
            genetic_test.kit_current_times += 1
            genetic_test.save()
            continue

        person_data = {}
        kits = GeneticTestKit.objects.filter(Q(genetic_test=genetic_test))
        if len(kits) == 0:
            first_kit = ""
            last_kit = ""
        elif len(kits) == 1:
            first_kit = kits.first().kit_code
            last_kit = ""
        elif len(kits) > 1:
            first_kit = kits.first().kit_code
            last_kit = kits.last().kit_code


        kit_status = genetic_test.kit_status
        if genetic_test.kit_current_times == genetic_test.total_times:
            billing_price = 3337
        else:
            billing_price = 3333

        if kit_status == "normal":
            pass
        if kit_status == "cancel":
            billing_price = 0
            genetic_test.is_active = False
        elif kit_status == "damaged_kit_return":
            billing_price = 20000
            genetic_test.is_active = False
        elif kit_status == "normal_kit_return":
            billing_price = 6000
            genetic_test.is_active = False


        user_id = genetic_test.customer.username
        SAT = get_sat()
        users, status = search_user(user_id, SAT)
        user = next(
            (item for item in users if item["username"] == user_id), None
        )
        attr_dict = {"name": "", "phone": "", "fruit_match": True}
        if user:
            if "attributes" in user:
                if "name" in user["attributes"]:
                    attr_dict["name"] = user["attributes"]["name"][0]
                if "phone" in user["attributes"]:
                    attr_dict["phone"] = user["attributes"]["phone"][0]
        else:
            return

        
        # person_objects = settlements.filter(genetic_promotion__coupon=coupon)
        # person_data = person_objects.aggregate(정산금액=Sum("billing_price"))
        # person_object = person_objects.last()
        # person_data["쿠폰번호"] = coupon
        person_data["매출 항목명"] = "할부" if genetic_test.kit_status == "normal" else genetic_test.get_kit_status_display()
        # person_data["거래처"] = person_object.genetic_promotion.promotion_product.connection_client
        # person_data["상품코드"] = person_object.genetic_promotion.promotion_product.product_code
        # person_data["상품명"] = person_object.genetic_promotion.promotion_product.product_name
        person_data["고객명"] = genetic_test.customer.name
        # person_data["아이디"] = person_object.genetic_promotion.customer_id
        person_data["이동전화번호"] = attr_dict["phone"]
        person_data["키트번호(최초)"] = first_kit
        person_data["키트번호(마지막 재채취)"] = last_kit
        # person_data["가입일"] = person_object.genetic_promotion.sign_datetime
        person_data["누적 회차"] = genetic_test.kit_current_times
        person_data["남은 회차"] = genetic_test.total_times - genetic_test.kit_current_times
        person_data["정산 금액"] = str(billing_price)
        # person_data["월별 누적"] = person_object.accumulate_times
        # person_data["해지일"] = person_object.genetic_promotion.termination_datetime if person_object.genetic_promotion.termination_datetime else ""
        person_data_list.append(person_data)


        GeneticTestLog.objects.create(
            order_date=datetime.datetime.now(),
            genetic_test=genetic_test,
            action="subscription",
            billing_price=billing_price,
            kit_current_times=genetic_test.kit_current_times,
        )

        genetic_test.kit_current_times += 1
        if genetic_test.kit_current_times == genetic_test.total_times:
            genetic_test.is_active = False
        genetic_test.save()

    logs = GeneticTestLog.objects.all()
    logs_qs = []
    max_kit_length = 0
    for item in logs:
        kits = item.genetic_test.genetictestkit_set.all()
        data = []
        data.append(item.order_date.date())
        data.append(item.order_date.strftime("%p %I:%M"))
        data.append(item.genetic_test.customer.name)
        data.append(item.genetic_test.customer.username)
        data.append(item.genetic_test.phone.as_national)
        data.append(item.genetic_test.test_product.get_connection_client_display())
        data.append(item.genetic_test.test_product.product_code)
        data.append(item.genetic_test.test_product.product_name)
        data.append(item.get_action_display())
        data.append(item.kit_current_times)

        for kit in kits:
            data.append(kit.kit_code)

        if max_kit_length < len(kits):
            max_kit_length = len(kits)

        logs_qs.append(data)
    
    columns=["날짜", "시간", "고객", "아이디", "전화번호", "거래처", "상품코드","상품명", "이벤트", "누적회차"]
    for ind in range(1, max_kit_length+1):
        columns.append(f"키트번호{ind}")
    logs_df = pd.DataFrame(logs_qs, columns=columns)

    user_ids = list(set(logs_df["아이디"].to_list()))
    new_df = pd.DataFrame(columns=columns)
    for user in user_ids:
        user_df = logs_df.groupby(["아이디"]).get_group(user)
        user_df = user_df.sort_values(["날짜", "시간"], axis=0)
        new_df = new_df.append(user_df)
        new_df = new_df.append(pd.Series(), ignore_index=True)
    new_df = new_df.fillna("")

    df = pd.DataFrame.from_dict(person_data_list, orient='columns')

    wb = openpyxl.load_workbook("data/macorgen_settlement_seet.xlsx")
    wb.save("data/tmp.xlsx")

    excel_writer = pd.ExcelWriter("data/tmp.xlsx", engine='openpyxl', mode='a')
    df.to_excel(excel_writer, sheet_name="월별현황", index=False)
    new_df.to_excel(excel_writer, sheet_name="거래 내역", index=False)
    excel_writer.save()

    wb = openpyxl.load_workbook("data/tmp.xlsx")
    wb.active
    for col in range(1, 12):
        wb["월별현황"].column_dimensions[get_column_letter(col)].width = 15
        wb["거래 내역"].column_dimensions[get_column_letter(col)].width = 15
        wb["거래 내역"].column_dimensions["H"].width = 40

    wb.save("data/tmp.xlsx")

    result = GeneticTestLogResult()
    today = datetime.datetime.now()

    with open('data/tmp.xlsx', 'rb') as excel:
            result.result_file_excel.save('data'+today.strftime('%Y-%m-%d %H:%M:%S')+'.xlsx', File(excel))
            result.save()


# 매달 말일 23:59:59에 스크립트 동작
class Command(BaseCommand):
    help = '마크로젠 정산'

    def handle(self, *args, **options):
        today = datetime.datetime.today()
        this_month_first = datetime.datetime(today.year, today.month, 1)
        next_month = datetime.datetime(today.year, today.month, 1) + relativedelta(months=1)
        this_month_last = next_month + relativedelta(seconds=-1)

        print('이번달 첫일: ' + this_month_first.strftime('%Y-%m-%d %H:%M:%S'))
        print('이번달 말일: ' + this_month_last.strftime('%Y-%m-%d %H:%M:%S'))

        first_datetime = this_month_first.strftime('%Y-%m-%d %H:%M:%S')
        last_datetime = this_month_last.strftime('%Y-%m-%d %H:%M:%S')
        last_date = this_month_last.strftime('%Y-%m-%d')
        
        # if today.strftime('%Y-%m-%d') == last_date:
        #     get_settlement(first_datetime, last_datetime, macrogen_data_list)
        get_settlement(first_datetime, last_datetime)

        # last_month_first = datetime(today.year, today.month, 1) + relativedelta(months=-1)
        # last_month_last = datetime(today.year, today.month, 1) + relativedelta(seconds=-1)

        # print('지난달 첫일: ' + last_month_first.strftime('%Y-%m-%d %H:%M:%S'))
        # print('지난달 말일: ' + last_month_last.strftime('%Y-%m-%d %H:%M:%S'))