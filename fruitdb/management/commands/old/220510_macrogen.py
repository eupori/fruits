import requests
import urllib
import json
import csv
import os
import django
import datetime
import pandas as pd


import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.core.files import File
from django.db.models import Sum
from django.db.models import Q
from fruitdb.views.api.auth import get_username_from_uat, get_sat, search_user
from fruitdb.models import *
from dateutil.relativedelta import relativedelta
from rest_framework import serializers
from io import BytesIO as IO


def get_settlement(first_datetime, last_datetime):

    today = datetime.datetime.strptime("2022-05-14", '%Y-%m-%d')

    first_date=datetime.datetime.strptime(first_datetime, '%Y-%m-%d %H:%M:%S')
    last_date=datetime.datetime.strptime(last_datetime, '%Y-%m-%d %H:%M:%S')

    cols = ["매출항목명", "고객명", "이동전화번호", "구분", "키트번호(최초)", "키트번호(마지막 재채취)", "가입일", "해지일", "진맛과 비고", "마크로젠 비고", "키트상태",  "키트발송일",  "키트반송일",  "분석시작일",  "분석완료일", "누적 회차", "남은 회차"]
    etc = ["키트", "분석"]
    date_cols = []
    start_date = datetime.datetime.strptime("2021-03-01", "%Y-%m-%d").date()
    end_date = today.date() - datetime.timedelta(days=1)

    loop_date = start_date
    while loop_date.strftime("%y%m") != end_date.strftime("%y%m"):
        cols.append(loop_date.strftime("%Y년 %m월 청구매출"))
        date_cols.append(loop_date.strftime("%Y-%m"))
        loop_date+= relativedelta(months=1)
    
    genetic_tests = GeneticTest.objects.all()

    person_data_list = []
    for genetic_test in genetic_tests:

        kits = GeneticTestKit.objects.filter(Q(genetic_test=genetic_test))
        if len(kits) == 0:
            first_kit = ""
            last_kit = ""
        elif len(kits) == 1:
            first_kit = kits.first().kit_code
            last_kit = ""
        elif len(kits) > 1:
            first_kit = kits.first().kit_code
            last_kit = kits.last().kit_code

        kit_status = genetic_test.kit_status
        if genetic_test.current_times == genetic_test.total_times:
            billing_price = 3348
        else:
            billing_price = 3332

        if kit_status == "normal":
            pass
        if kit_status == "cancel":
            billing_price = 0
            genetic_test.is_active = False
        elif kit_status == "damaged_kit_return":
            billing_price = 20000
            genetic_test.is_active = False
        elif kit_status == "normal_kit_return":
            billing_price = 6000
            genetic_test.is_active = False

        ###################################
        if genetic_test.start_datetime + relativedelta(months=6) < today and not genetic_test.is_kit_recept and genetic_test.is_active:
            billing_price = 8
            genetic_test.is_active=False
            genetic_test.end_datetime = today

            GeneticTestLog.objects.create(
                order_date=today,
                genetic_test=genetic_test,
                action="subscription",
                billing_price=billing_price,
                current_times=genetic_test.current_times,
            )
        if genetic_test.status == "cancel" and genetic_test.is_active:
            if genetic_test.is_kit_recept:
                billing_price = 40000 - (genetic_test.current_times * 3332)
                genetic_test.kit_status = "cancel"
            if genetic_test.end_datetime > genetic_test.start_datetime + relativedelta(days=14):
                billing_price = 40000 - (genetic_test.current_times * 3332)
                genetic_test.kit_status = "cancel"
            else:
                if genetic_test.kit_status == "normal_kit_return":
                    billing_price = 6000
                elif genetic_test.kit_status == "damaged_kit_return":
                    billing_price = 20000
            genetic_test.is_active = False
            genetic_test.end_datetime = today

            GeneticTestLog.objects.create(
                order_date=today,
                genetic_test=genetic_test,
                action="subscription",
                billing_price=billing_price,
                current_times=genetic_test.current_times,
            )

        ###################################

        if genetic_test.is_active:
            GeneticTestLog.objects.create(
                order_date=today,
                genetic_test=genetic_test,
                action="subscription",
                billing_price=billing_price,
                current_times=genetic_test.current_times,
            )

        for each in etc:
            person_data = {}
            for col in cols:
                person_data[col] = "-"
            person_data["매출항목명"] = "할부" if genetic_test.kit_status == "normal" else "해지"
            # person_data["이동전화번호"] = attr_dict["phone"]
            person_data["고객명"] = genetic_test.customer.name
            person_data["이동전화번호"] = genetic_test.phone.as_national
            person_data["구분"] = each
            person_data["키트번호(최초)"] = first_kit
            person_data["키트번호(마지막 재채취)"] = last_kit
            person_data["가입일"] = genetic_test.start_datetime.strftime("%Y.%m.%d")
            person_data["해지일"] = genetic_test.end_datetime.strftime("%Y.%m.%d") if genetic_test.end_datetime else "-"
            person_data["진맛과 비고"] = ""
            person_data["마크로젠 비고"] = ""
            person_data["키트상태"] = ""
            person_data["키트발송일"] = ""
            person_data["키트반송일"] = ""
            person_data["분석시작일"] = ""
            person_data["분석완료일"] = ""
            person_data["누적 회차"] = genetic_test.current_times
            person_data["남은 회차"] = genetic_test.total_times - genetic_test.current_times

            action_list =  ["normal_kit_return", "damaged_kit_return", "subscription_cancel", "subscription"]
            person_logs = genetic_test.genetictestlog_set.filter(action__in=action_list)
            for pl in person_logs:
                person_data[f'{pl.order_date.strftime("%Y")}년 {pl.order_date.strftime("%m")}월 청구매출'] = pl.billing_price/2

            person_data_list.append(person_data)

        if genetic_test.current_times == genetic_test.total_times:
            genetic_test.is_active = False
            genetic_test.end_datetime = today
        else:
            genetic_test.current_times += 1 if genetic_test.is_active else 0
        genetic_test.save()

    logs = GeneticTestLog.objects.all()
    logs_qs = []
    max_kit_length = 0
    for item in logs:
        kits = item.genetic_test.genetictestkit_set.all()
        data = []
        data.append(item.order_date.date())
        data.append(item.order_date.strftime("%p %I:%M"))
        data.append(item.genetic_test.customer.name)
        data.append(item.genetic_test.customer.username)
        data.append(item.genetic_test.phone.as_national)
        data.append(item.genetic_test.test_product.get_connection_client_display())
        data.append(item.genetic_test.test_product.product_code)
        data.append(item.genetic_test.test_product.product_name)
        data.append(item.get_action_display())
        data.append(item.current_times)

        for kit in kits:
            data.append(kit.kit_code)

        if max_kit_length < len(kits):
            max_kit_length = len(kits)

        logs_qs.append(data)
    
    columns=["날짜", "시간", "고객", "아이디", "전화번호", "거래처", "상품코드","상품명", "이벤트", "누적회차"]
    for ind in range(1, max_kit_length+1):
        columns.append(f"키트번호{ind}")
    logs_df = pd.DataFrame(logs_qs, columns=columns)

    user_ids = list(set(logs_df["아이디"].to_list()))
    new_df = pd.DataFrame(columns=columns)
    for user in user_ids:
        user_df = logs_df.groupby(["아이디"]).get_group(user)
        user_df = user_df.sort_values(["날짜", "시간"], axis=0)
        new_df = new_df.append(user_df)
        new_df = new_df.append(pd.Series(), ignore_index=True)
    new_df = new_df.fillna("")

    df = pd.DataFrame.from_dict(person_data_list, orient='columns')

    wb = openpyxl.load_workbook("data/macorgen_settlement_seet.xlsx")
    wb.save("data/tmp.xlsx")

    excel_writer = pd.ExcelWriter("data/tmp.xlsx", engine='openpyxl', mode='a')
    df.to_excel(excel_writer, sheet_name="월별현황", index=False)
    new_df.to_excel(excel_writer, sheet_name="거래 내역", index=False)
    excel_writer.save()

    wb = openpyxl.load_workbook("data/tmp.xlsx")
    wb.active
    for col in range(1, 16):
        wb["월별현황"].column_dimensions[get_column_letter(col)].width = 15
        wb["거래 내역"].column_dimensions[get_column_letter(col)].width = 15
        wb["거래 내역"].column_dimensions["H"].width = 40

    wb.save("data/tmp.xlsx")

    result = GeneticTestLogResult()
    today = datetime.datetime.now()

    with open('data/tmp.xlsx', 'rb') as excel:
            result.result_file_excel.save('macorgen_settlement_'+today.strftime('%Y-%m-%d %H:%M:%S')+'.xlsx', File(excel))
            result.save()


# 매달 말일 23:59:59에 스크립트 동작
class Command(BaseCommand):
    help = '마크로젠 정산'

    def handle(self, *args, **options):
        today = datetime.datetime.today()
        this_month_first = datetime.datetime(today.year, today.month, 1)
        next_month = datetime.datetime(today.year, today.month, 1) + relativedelta(months=1)
        this_month_last = next_month + relativedelta(seconds=-1)

        print('이번달 첫일: ' + this_month_first.strftime('%Y-%m-%d %H:%M:%S'))
        print('이번달 말일: ' + this_month_last.strftime('%Y-%m-%d %H:%M:%S'))

        first_datetime = this_month_first.strftime('%Y-%m-%d %H:%M:%S')
        last_datetime = this_month_last.strftime('%Y-%m-%d %H:%M:%S')
        last_date = this_month_last.strftime('%Y-%m-%d')
        
        # if today.strftime('%Y-%m-%d') == last_date:
        #     get_settlement(first_datetime, last_datetime, macrogen_data_list)
        get_settlement(first_datetime, last_datetime)

        # last_month_first = datetime(today.year, today.month, 1) + relativedelta(months=-1)
        # last_month_last = datetime(today.year, today.month, 1) + relativedelta(seconds=-1)

        # print('지난달 첫일: ' + last_month_first.strftime('%Y-%m-%d %H:%M:%S'))
        # print('지난달 말일: ' + last_month_last.strftime('%Y-%m-%d %H:%M:%S'))